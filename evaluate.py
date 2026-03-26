#!/usr/bin/env python3
"""
evaluate.py — TRUE OUT-OF-SAMPLE evaluation + full trade log.

Outputs two CSV files to logs/:
  trades_OOS_YYYYMMDD.csv   — every single trade with full detail
  daily_OOS_YYYYMMDD.csv    — daily portfolio snapshot

Run once. Do NOT use results to tune agent.py — contaminates the test.
"""
import sys, signal, csv
from pathlib import Path
from datetime import datetime
import numpy as np
import pandas as pd

signal.signal(signal.SIGALRM, signal.SIG_DFL)
signal.alarm(0)

from prepare import load_data
from agent import generate_signals, HYPOTHESIS

STARTING_CAPITAL     = 100_000.0
COMMISSION_PER_TRADE = 20.0
SLIPPAGE_BPS         = 5

LOG_DIR = Path("logs")
LOG_DIR.mkdir(exist_ok=True)


def run_oos_backtest(weights: pd.DataFrame, data: dict) -> tuple:
    """
    Returns: (metrics_dict, trade_log, daily_log)
    trade_log  — one dict per trade executed
    daily_log  — one dict per trading day
    """
    close  = data["close"]
    open_  = data["open"]
    idx    = close.index

    oos_candidates = idx[idx >= pd.Timestamp("2024-07-01")]
    if len(oos_candidates) == 0:
        print("[ERROR] No data after 2024-07-01. Run refresh_data.py.")
        sys.exit(1)
    oos_start = oos_candidates[0]
    print(f"  OOS start : {oos_start.date()} | end: {idx[-1].date()} | days: {len(oos_candidates)}")

    close_oos = close.loc[oos_start:]
    open_oos  = open_.loc[oos_start:]
    n_days    = len(close_oos)

    if n_days < 20:
        print(f"\n[ERROR] OOS too short ({n_days} days). Run refresh_data.py first.")
        sys.exit(1)

    weights_oos = (
        weights
        .reindex(close_oos.index)
        .reindex(columns=close_oos.columns)
        .fillna(0.0)
    )

    slip         = SLIPPAGE_BPS / 10_000.0
    cash         = STARTING_CAPITAL
    shares_held  = pd.Series(0.0, index=close_oos.columns)
    prev_w_norm  = pd.Series(0.0, index=close_oos.columns)
    port_vals    = []
    comm_total   = slip_total = trade_count = 0.0

    trade_log = []
    daily_log = []

    for i in range(n_days):
        date      = close_oos.index[i]
        px_open   = open_oos.iloc[i]
        px_close_prev = close_oos.iloc[i - 1] if i > 0 else px_open

        port_value = cash + (shares_held * px_open).sum()
        port_vals.append(port_value)

        # Daily snapshot
        held = shares_held[shares_held > 0.001]
        daily_log.append({
            "date"            : str(date.date()),
            "portfolio_value" : round(port_value, 2),
            "cash"            : round(cash, 2),
            "n_positions"     : len(held),
            "positions"       : " | ".join(
                f"{t}:{shares_held[t]:.1f}sh@{px_open.get(t,0):.2f}"
                for t in held.index
            ),
        })

        if i == 0:
            continue

        target_w = weights_oos.iloc[i - 1]
        gross    = target_w.abs().sum()
        w_norm   = target_w / gross if gross > 0 else target_w

        target_shares = ((w_norm * port_value) / px_open).fillna(0.0)
        target_shares = target_shares.where(px_open > 0, 0.0)

        changed      = (w_norm - prev_w_norm).abs() > 1e-6
        delta        = (target_shares - shares_held).where(changed, 0.0)
        traded       = delta[delta.abs() > 0.01].index

        trade_cash = 0.0
        batch = []
        for tkr in traded:
            d  = delta[tkr]
            px = px_open[tkr]
            if px <= 0 or np.isnan(px):
                continue

            if d > 0:
                fill_px = px * (1 + slip)
                action  = "BUY"
            else:
                fill_px = px * (1 - slip)
                action  = "SELL"

            gross_val  = abs(d) * fill_px
            slip_cost  = abs(d) * px * slip
            comm_cost  = COMMISSION_PER_TRADE

            trade_cash  -= d * fill_px
            slip_total  += slip_cost
            comm_total  += comm_cost
            trade_count += 1

            was_held  = shares_held[tkr] > 0.001
            will_hold = target_shares[tkr] > 0.001
            if action == "BUY" and not was_held:
                reason = "REBAL_ENTER"
            elif action == "SELL" and not will_hold and changed.get(tkr, False):
                reason = "REBAL_EXIT"
            elif action == "SELL" and not will_hold:
                reason = "STOP_EXIT"
            else:
                reason = "REBAL_RESIZE"

            signal_px = float(px_close_prev.get(tkr, px))
            pos_pct   = gross_val / port_value * 100

            batch.append({
                "exec_date"          : str(date.date()),
                "signal_date"        : str(close_oos.index[i-1].date()),
                "ticker"             : tkr,
                "action"             : action,
                "reason"             : reason,
                "shares"             : round(abs(d), 4),
                "fill_price"         : round(fill_px, 4),
                "signal_close"       : round(signal_px, 4),
                "slippage_bps"       : SLIPPAGE_BPS,
                "slippage_$"         : round(slip_cost, 2),
                "commission_$"       : comm_cost,
                "gross_value_$"      : round(gross_val, 2),
                "portfolio_$_before" : round(port_value, 2),
                "position_pct"       : round(pos_pct, 2),
                "cash_after_$"       : None,
            })

        shares_held = target_shares.copy()
        prev_w_norm = w_norm.copy()
        cash       += trade_cash - COMMISSION_PER_TRADE * len(traded)

        for row in batch:
            row["cash_after_$"] = round(cash, 2)
        trade_log.extend(batch)

    pv  = pd.Series(port_vals, index=close_oos.index)
    r   = pv.pct_change().fillna(0.0)
    ann = 252

    if "SPY" in open_oos.columns:
        br = (open_oos["SPY"].shift(-1) / open_oos["SPY"] - 1).fillna(0.0)
    else:
        br = (open_oos.shift(-1) / open_oos - 1).fillna(0.0).mean(axis=1)

    bench_pv = STARTING_CAPITAL * (1 + br).cumprod()
    tot      = pv.iloc[-1] / STARTING_CAPITAL - 1
    btot     = bench_pv.iloc[-1] / STARTING_CAPITAL - 1
    tpy      = trade_count / (n_days / ann)

    def sharpe(r): return (r.mean()/r.std()*np.sqrt(ann)) if r.std()>1e-9 else 0.0
    def calmar(r):
        cum=(1+r).cumprod(); dd=(cum/cum.cummax()-1).min()
        ar=cum.iloc[-1]**(ann/max(len(r),1))-1
        return ar/abs(dd) if dd<-1e-6 else 0.0
    def maxdd(r): cum=(1+r).cumprod(); return (cum/cum.cummax()-1).min()

    rebal_count = sum(1 for t in trade_log if "REBAL" in t["reason"])
    stop_count  = sum(1 for t in trade_log if t["reason"] == "STOP_EXIT")

    metrics = dict(
        oos_start        = str(oos_start.date()),
        oos_end          = str(close_oos.index[-1].date()),
        n_days           = n_days,
        sharpe           = round(sharpe(r), 3),
        calmar           = round(calmar(r), 3),
        max_drawdown     = round(maxdd(r)*100, 2),
        win_rate         = round((r>0).mean()*100, 1),
        ann_vol          = round(r.std()*np.sqrt(ann)*100, 2),
        total_return_pct = round(tot*100, 2),
        bench_return_pct = round(btot*100, 2),
        alpha_ann        = round((r-br).mean()*ann*100, 2),
        final_value      = round(pv.iloc[-1], 0),
        bench_final      = round(bench_pv.iloc[-1], 0),
        net_pnl          = round(pv.iloc[-1]-STARTING_CAPITAL, 0),
        total_trades     = int(trade_count),
        trades_per_year  = round(tpy, 1),
        rebal_trades     = rebal_count,
        stop_trades      = stop_count,
        commission_paid  = round(comm_total, 0),
        slippage_paid    = round(slip_total, 0),
        total_cost       = round(comm_total+slip_total, 0),
        cost_pct_capital = round((comm_total+slip_total)/STARTING_CAPITAL*100, 2),
    )
    return metrics, trade_log, daily_log


def save_logs(trade_log: list, daily_log: list, tag: str, metrics: dict, hypothesis: str):
    xlsx_file  = LOG_DIR / f"OOS_report_{tag}.xlsx"
    trade_file = LOG_DIR / f"trades_OOS_{tag}.csv"
    daily_file = LOG_DIR / f"daily_OOS_{tag}.csv"

    # ── CSV (backup) ──────────────────────────────────────────────────────────
    if trade_log:
        with open(trade_file, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(trade_log[0].keys()))
            w.writeheader(); w.writerows(trade_log)

    if daily_log:
        with open(daily_file, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(daily_log[0].keys()))
            w.writeheader(); w.writerows(daily_log)

    # ── Excel workbook ────────────────────────────────────────────────────────
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Helper styles ─────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
    HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
    TITLE_FONT = Font(bold=True, size=12)
    MONO_FONT  = Font(name="Courier New", size=9)
    GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
    RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
    GREY_FILL  = PatternFill("solid", fgColor="F2F2F2")
    BORDER     = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header_row(ws, row_num, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

    def autofit(ws, min_w=8, max_w=40):
        for col in ws.columns:
            length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1: Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws["B2"] = "OOS Evaluation Report"
    ws["B2"].font = Font(bold=True, size=14, color="1F3864")
    ws["B3"] = hypothesis
    ws["B3"].font = Font(italic=True, size=10, color="595959")
    ws["B4"] = f"Period: {metrics['oos_start']} → {metrics['oos_end']}  ({metrics['n_days']} trading days)"

    sections = [
        ("Performance", [
            ("Sharpe Ratio",       metrics["sharpe"]),
            ("Calmar Ratio",       metrics["calmar"]),
            ("Max Drawdown",       f"{metrics['max_drawdown']}%"),
            ("Win Rate (daily)",   f"{metrics['win_rate']}%"),
            ("Ann. Volatility",    f"{metrics['ann_vol']}%"),
        ]),
        ("Returns", [
            ("Strategy Return",    f"+{metrics['total_return_pct']}%"),
            ("SPY Return",         f"+{metrics['bench_return_pct']}%"),
            ("Alpha (ann.)",       f"{metrics['alpha_ann']}%"),
        ]),
        ("Dollar P&L", [
            ("Starting Capital",   f"${STARTING_CAPITAL:,.0f}"),
            ("Final Value",        f"${metrics['final_value']:,.0f}"),
            ("Net P&L",            f"${metrics['net_pnl']:,.0f}"),
            ("SPY Final Value",    f"${metrics['bench_final']:,.0f}"),
        ]),
        ("Trading Costs", [
            ("Total Trades",       metrics["total_trades"]),
            ("Trades / Year",      metrics["trades_per_year"]),
            ("Rebal Trades",       metrics["rebal_trades"]),
            ("Stop Exits",         metrics["stop_trades"]),
            ("Commission Paid",    f"${metrics['commission_paid']:,.0f}"),
            ("Slippage Paid",      f"${metrics['slippage_paid']:,.0f}"),
            ("Total Cost",         f"${metrics['total_cost']:,.0f}  ({metrics['cost_pct_capital']}% of capital)"),
        ]),
    ]

    row = 6
    for section_title, rows in sections:
        ws.cell(row=row, column=2).value = section_title
        ws.cell(row=row, column=2).font = Font(bold=True, size=11, color="1F3864")
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.merge_cells(f"B{row}:C{row}")
        row += 1
        for label, value in rows:
            lc = ws.cell(row=row, column=2, value=label)
            vc = ws.cell(row=row, column=3, value=value)
            lc.font = Font(size=10, bold=True)
            vc.font = Font(size=10)
            if row % 2 == 0:
                lc.fill = GREY_FILL
                vc.fill = GREY_FILL
            lc.border = BORDER; vc.border = BORDER
            row += 1
        row += 1

    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2: Trades
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Trades")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A2"

    if trade_log:
        headers = list(trade_log[0].keys())
        for c, h in enumerate(headers, 1):
            ws2.cell(row=1, column=c, value=h)
        style_header_row(ws2, 1, len(headers))

        for r, trade in enumerate(trade_log, 2):
            for c, key in enumerate(headers, 1):
                cell = ws2.cell(row=r, column=c, value=trade[key])
                cell.border = BORDER
                cell.font = Font(size=9)
                # Colour rows by action
                if trade["action"] == "BUY":
                    cell.fill = PatternFill("solid", fgColor="EBF5FB")
                elif trade["reason"] == "STOP_EXIT":
                    cell.fill = RED_FILL
                elif trade["action"] == "SELL":
                    cell.fill = PatternFill("solid", fgColor="FDFEFE")
                if r % 2 == 0 and trade["action"] not in ("BUY",) and trade["reason"] != "STOP_EXIT":
                    cell.fill = GREY_FILL

        autofit(ws2)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 3: Daily
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Daily")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A2"

    if daily_log:
        headers = list(daily_log[0].keys())
        for c, h in enumerate(headers, 1):
            ws3.cell(row=1, column=c, value=h)
        style_header_row(ws3, 1, len(headers))

        prev_val = STARTING_CAPITAL
        for r, day in enumerate(daily_log, 2):
            for c, key in enumerate(headers, 1):
                cell = ws3.cell(row=r, column=c, value=day[key])
                cell.border = BORDER
                cell.font = Font(size=9)
                if r % 2 == 0:
                    cell.fill = GREY_FILL
            # Colour portfolio value cell green/red vs previous day
            pv_cell = ws3.cell(row=r, column=2)
            cur_val = day["portfolio_value"]
            if isinstance(cur_val, (int, float)):
                pv_cell.fill = GREEN_FILL if cur_val >= prev_val else RED_FILL
                prev_val = cur_val

        # Wide column for positions string
        ws3.column_dimensions["E"].width = 80
        autofit(ws3, max_w=20)
        ws3.column_dimensions["E"].width = 80

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 4: Holdings snapshot (current open positions at end of OOS)
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Holdings")
    ws4.sheet_view.showGridLines = False

    ws4["A1"] = "Open positions at end of OOS period"
    ws4["A1"].font = Font(bold=True, size=11, color="1F3864")

    if daily_log:
        last_day = daily_log[-1]
        ws4["A2"] = f"Date: {last_day['date']}  |  Portfolio: ${last_day['portfolio_value']:,.2f}  |  Cash: ${last_day['cash']:,.2f}"
        ws4["A2"].font = Font(size=10, italic=True)

        headers = ["Ticker", "Shares", "Last Price", "Market Value", "% Portfolio"]
        for c, h in enumerate(headers, 1):
            ws4.cell(row=4, column=c, value=h)
        style_header_row(ws4, 4, len(headers))

        positions_str = last_day.get("positions", "")
        row = 5
        if positions_str:
            for pos in positions_str.split(" | "):
                if not pos.strip():
                    continue
                try:
                    tkr_part, rest = pos.split(":")
                    sh_part, px_part = rest.split("sh@")
                    shares = float(sh_part)
                    price  = float(px_part)
                    mkt_val = shares * price
                    pct     = mkt_val / last_day["portfolio_value"] * 100
                    for c, val in enumerate([tkr_part, round(shares,2), round(price,4),
                                             round(mkt_val,2), round(pct,2)], 1):
                        cell = ws4.cell(row=row, column=c, value=val)
                        cell.border = BORDER
                        cell.font = Font(size=10)
                        if row % 2 == 0:
                            cell.fill = GREY_FILL
                    row += 1
                except Exception:
                    continue

    autofit(ws4)

    wb.save(xlsx_file)
    print(f"  Excel     : {xlsx_file}  (4 sheets: Summary, Trades, Daily, Holdings)")
    print(f"  CSV trade : {trade_file}  ({len(trade_log)} rows)")
    print(f"  CSV daily : {daily_file}  ({len(daily_log)} rows)")


def print_report(m, hypothesis, trade_log):
    W = 66
    print("\n" + "="*W)
    print("  TRUE OUT-OF-SAMPLE EVALUATION")
    print("  This data was NEVER used during training or experiment selection.")
    print("="*W)
    print(f"  Strategy   : {hypothesis}")
    print(f"  OOS period : {m['oos_start']} → {m['oos_end']}  ({m['n_days']} trading days)")
    print(f"  Capital    : ${STARTING_CAPITAL:,.0f}  |  "
          f"Commission: ${COMMISSION_PER_TRADE:.0f}/trade  |  Slippage: {SLIPPAGE_BPS}bps")
    print("-"*W)
    rows = [
        ("",                 "── Core ──────────────────────────────"),
        ("Sharpe",           m["sharpe"]),
        ("Calmar",           m["calmar"]),
        ("Max drawdown",     f"{m['max_drawdown']}%"),
        ("Win rate",         f"{m['win_rate']}%"),
        ("Ann. volatility",  f"{m['ann_vol']}%"),
        ("",                 "── Returns ───────────────────────────"),
        ("Strategy return",  f"+{m['total_return_pct']}%"),
        ("SPY return",       f"+{m['bench_return_pct']}%"),
        ("Alpha (ann.)",     f"{m['alpha_ann']}%"),
        ("",                 "── Dollar P&L ────────────────────────"),
        ("Starting capital", f"${STARTING_CAPITAL:,.0f}"),
        ("Final value",      f"${m['final_value']:,.0f}"),
        ("Net P&L",          f"${m['net_pnl']:,.0f}"),
        ("SPY final",        f"${m['bench_final']:,.0f}"),
        ("",                 "── Costs ─────────────────────────────"),
        ("Total trades",     m["total_trades"]),
        ("Trades/year",      m["trades_per_year"]),
        ("  Rebal trades",   m["rebal_trades"]),
        ("  Stop exits",     m["stop_trades"]),
        ("Commission",       f"${m['commission_paid']:,.0f}"),
        ("Slippage",         f"${m['slippage_paid']:,.0f}"),
        ("Total cost",       f"${m['total_cost']:,.0f}  ({m['cost_pct_capital']}% of capital)"),
    ]
    for label, value in rows:
        if label == "":
            print(f"\n  {value}")
        else:
            print(f"    {label:<22}: {value}")

    if trade_log:
        from collections import Counter
        reasons = Counter(t["reason"] for t in trade_log)
        print(f"\n  [Trade breakdown by reason]")
        for reason, count in sorted(reasons.items()):
            print(f"    {reason:<20}: {count}")

        # Best/worst completed round-trips
        buys = {}
        for t in trade_log:
            if t["action"] == "BUY" and t["ticker"] not in buys:
                buys[t["ticker"]] = t["fill_price"]
        completed = {}
        for t in trade_log:
            if t["action"] == "SELL" and t["ticker"] in buys:
                pnl = (t["fill_price"] / buys[t["ticker"]] - 1) * 100
                completed[t["ticker"]] = (pnl, t["reason"])

        if completed:
            srt = sorted(completed.items(), key=lambda x: x[1][0])
            print(f"\n  [Round-trips — worst 5]")
            for tkr, (pnl, rsn) in srt[:5]:
                print(f"    {tkr:<8}: {pnl:+.1f}%  ({rsn})")
            print(f"  [Round-trips — best 5]")
            for tkr, (pnl, rsn) in srt[-5:]:
                print(f"    {tkr:<8}: {pnl:+.1f}%  ({rsn})")

    s    = m["sharpe"]
    beat = m["total_return_pct"] > m["bench_return_pct"]
    verdict = (
        "STRONG  — Sharpe >= 1.5. Generalised well. Live trading candidate." if s >= 1.5 else
        "GOOD    — Sharpe >= 1.0. Solid out-of-sample performance."          if s >= 1.0 else
        "WEAK    — Sharpe >= 0.5. Thin edge. Paper trade first."             if s >= 0.5 else
        "FAILED  — Sharpe < 0.5. Did not generalise."
    )
    print(f"\n{'='*W}")
    print(f"  Verdict  : {verdict}")
    print(f"  Beat SPY : {'YES' if beat else 'NO'} "
          f"({m['total_return_pct']}% vs {m['bench_return_pct']}%)")
    print("="*W + "\n")


if __name__ == "__main__":
    print("\n>>> Loading data and generating signals …")
    data    = load_data()
    weights = generate_signals(data)

    print("\n>>> Running true out-of-sample backtest …")
    metrics, trade_log, daily_log = run_oos_backtest(weights, data)

    print("\n>>> Saving logs …")
    tag = datetime.today().strftime("%Y%m%d")
    save_logs(trade_log, daily_log, tag, metrics, HYPOTHESIS)

    print_report(metrics, HYPOTHESIS, trade_log)
